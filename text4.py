import pandas as pd
from Bio import SeqIO           #SeqIO用于更为快捷的得到KMER
from collections import Counter   #用于统计可哈希对象的出现次数
import itertools
from openpyxl import load_workbook  #用于向EXCEL中写入东西(追加)

path=f"C:/Users/ADMIN/Desktop/1000.fasta"        #已下载好的基因组信息
path2=f"C:/Users/ADMIN/Desktop/text2.txt"      #新建的存放信息的文件
path3="C:/Users/ADMIN/Desktop/1000.tsv"         #不同样本的CLADE信息
path4="C:/Users/ADMIN/Desktop/text3.txt"        #存放每个样本的名字
path5="D:/Pycharm_project/project_2/新冠病毒亚型统计.xlsx"   #存放不同亚型的样本名（1000中三种类型各选64个）
path_a="C:/Users/ADMIN/Desktop/delta.txt"
path_b="C:/Users/ADMIN/Desktop/omicron.txt"
path_c="C:/Users/ADMIN/Desktop/alpha.txt"
path_6="C:/Users/ADMIN/Desktop/Kmer频数表.xlsx"   #用于计算卡方值的表格


#依据WHO文件规定的亚型鉴别规则
Delta=['Delta','21A','21I','21J']
Omicron=['Omicron','21M','21L','21K','22A','22B','22C']
Alpha=['Alpha','20I']
Gamma=['Gamma','21J']
Beta=['Beta','20H']
Epsilon=['Epsilon','21C']
clade=[Delta,Omicron,Alpha]     #实际上，除了这三种外，其他类亚型基因组数量很少，故先排除

#该函数用于将样本的名字写入text3
def write_to_name(sample_name):
    with open(path4,'a')as f:
        f.write(sample_name)
        f.write("\n")
#该函数用于删除text2和text4中原本的内容
def delete_txt():
    with open(path2, 'w') as f1:
        f1.write("")
    with open(path4,'w')as f2:
        f2.write("")
#该函数用于将下载的基因组信息重写入text2（原本的含一些杂项）
def write_to_txt():
    df = pd.read_excel(path5)
    c = df["total"].tolist()
    with open(path,'r',encoding="utf-8-sig")as f:
        lines=f.readlines()
        sign = 0
        for line in lines:
            #line=line.strip()    发现在该项目中，>被识别为空白字符，原因不明，待解决
            if line[0]=='>'and not (line[1:11] in c):
                sign=0             #这段代码写的有点愚蠢，原理大概是sign起到一个信号灯的作用。
            if line[0]=='>'and line[1:11] in c:
                with open(path2,'a')as t1:
                    t1.write("\n")               #这样的代码会在第一行留出一个空行，需要手动删掉
                    t1.write(line[:11])
                    write_to_name(line[1:11])      #将样本名字信息写入text3
                    t1.write("\n")
                    sign=1
            elif sign==1:
                with open(path2,'a')as t2:
                    line = line.strip()
                    t2.write(line)
def write_to_doa(name):
    df=pd.read_excel(path5)
    c1=df["delta"].tolist()
    c2 = df["omicron"].tolist()
    c3 = df["alpha"].tolist()
    if name in c1:
        with open("C:/Users/ADMIN/Desktop/delta.txt",'a')as f:
            f.write(name)
            f.write("\n")
    elif name in c2:
        with open("C:/Users/ADMIN/Desktop/omicron.txt",'a')as f:
            f.write(name)
            f.write("\n")
    elif name in c3:
        with open("C:/Users/ADMIN/Desktop/alpha.txt",'a')as f:
            f.write(name)
            f.write("\n")
    else:
        print("未知分型")
#该函数用于返回特定样本的亚型。传入样本名，返回亚型
def identify_clade(sample_name):
    with open(path3,'r')as f:
        lines=f.readlines()
        for l in lines:
            l=l.strip()
            if l[0:10]==sample_name:
                for i in clade:
                    if l[-3:] in i:
                        return(i[0])
                return "others"
#该函数用于初始化哈希表
def generate_kmers(k):
    bases = 'ACGT'
    all_combinations = itertools.product(bases, repeat=k)
    all_kmers = {''.join(kmer): 0 for kmer in all_combinations}
    return all_kmers
#该函数用于计算得到KMER
def get_Kmer(path,path2, k=6):
    temp=[]
    with open(path2,'r')as f:
        lines=f.readlines()
        for i in lines:
            i=i.strip()
            temp.append(i)
    Kmers = generate_kmers(k)
    from Bio import SeqIO
    num=0
    for record in SeqIO.parse(path, "fasta"):
        seq = str(record.seq).upper()
        if str(record.id)in temp:
            num += 1
            for i in range(len(seq) - k + 1):
                kmer = seq[i:i + k]
                if kmer in Kmers:
                    Kmers[kmer] += 1
    print(f"本次共对{num}个对象进行了Kmer值计算")
    return Kmers
#该函数用于向EXCEL表格中写入用于计算卡方值的数据
def write_to_excel(dict1,n):  #n为预计写入第n+1行
    wb=load_workbook(path_6)
    ws=wb.active
    r=2
    c1=1
    c2=n+1
    for i in dict1.keys():
        ws.cell(row=r,column=c1,value=i)
        ws.cell(row=r, column=c2, value=dict1[i])
        r+=1
    wb.save(path_6)
#用于清空excel
def delete_excel():
    wb=load_workbook(path_6)
    ws = wb["Sheet"]
    ws.delete_rows(2, ws.max_row - 1)   #清空后第一行的内容要手动补充
    wb.save(path_6)
#用于同时计算多种亚型的Kmer
def get_three_Kmer(path1,path2,path3):
    echo=1
    for i in [path1,path2,path3]:
        Kmer=get_Kmer("C:/Users/ADMIN/Desktop/text2.fasta",i)
        write_to_excel(Kmer,echo)
        echo+=1
if __name__=="__main__":
    #delete_txt()
    #write_to_txt()
    """with open(path4,'r')as f:
        lines=f.readlines()
        for i in lines:
            i=i.strip()
            print(identify_clade(i))
    """
    """
    with open(path4,'r')as f:
        lines=f.readlines()
        for i in lines:
            i=i.strip()
            write_to_doa(i)
    """
    get_three_Kmer(path_a,path_b,path_c)
    print("\nover")